"""
modules/excel_renderer.py
Renders an Excel sheet to a PIL Image and computes cell bounding boxes
for the eye-popup cell-highlight feature.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw

from modules.cell_format import _resolve_color, format_cell_value_with_fmt


def _col_px(ws, c: int, scale: float = 1.0) -> int:
    letter = get_column_letter(c)
    cd     = ws.column_dimensions.get(letter)
    w      = cd.width if (cd and cd.width and cd.width > 0) else 8.43
    return max(60, int(w * 10 * scale))


def _row_px(ws, r: int, scale: float = 1.0) -> int:
    rd = ws.row_dimensions.get(r)
    h  = rd.height if (rd and rd.height and rd.height > 0) else 15.0
    return max(14, int(h * 1.5 * scale))


def render_excel_sheet(excel_path: str, sheet_name: str, scale: float = 1.0):
    """Returns (PIL Image, col_starts, row_starts, merged_master)."""
    wb      = openpyxl.load_workbook(excel_path, data_only=True)
    ws      = wb[sheet_name]
    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    col_starts = [0]
    for c in range(1, max_col + 1):
        col_starts.append(col_starts[-1] + _col_px(ws, c, scale))
    row_starts = [0]
    for r in range(1, max_row + 1):
        row_starts.append(row_starts[-1] + _row_px(ws, r, scale))

    img  = Image.new("RGB", (col_starts[-1], row_starts[-1]), "white")
    draw = ImageDraw.Draw(img, "RGBA")

    merged_master: dict = {}
    for mr in ws.merged_cells.ranges:
        mn_r, mn_c, mx_r, mx_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        for rr in range(mn_r, mx_r + 1):
            for cc in range(mn_c, mx_c + 1):
                merged_master[(rr, cc)] = (mn_r, mn_c, mx_r, mx_c)

    drawn_merges: set = set()
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            merge_info = merged_master.get((r, c))
            if merge_info:
                mn_r, mn_c, mx_r, mx_c = merge_info
                if (mn_r, mn_c) in drawn_merges:
                    continue
                drawn_merges.add((mn_r, mn_c))
                x1, y1 = col_starts[mn_c - 1], row_starts[mn_r - 1]
                x2, y2 = col_starts[mx_c],      row_starts[mx_r]
                cell   = ws.cell(mn_r, mn_c)
            else:
                x1, y1 = col_starts[c - 1], row_starts[r - 1]
                x2, y2 = col_starts[c],      row_starts[r]
                cell   = ws.cell(r, c)

            bg_hex = "FFFFFF"
            if cell.fill and cell.fill.fill_type == "solid":
                bg_hex = _resolve_color(cell.fill.fgColor, "FFFFFF")
            draw.rectangle([x1, y1, x2 - 1, y2 - 1], fill=f"#{bg_hex}", outline="#CCCCCC", width=1)

            val = cell.value
            if val is not None:
                txt_color = "#000000"
                if cell.font and cell.font.color:
                    fc = _resolve_color(cell.font.color, "000000")
                    if fc.upper() != bg_hex.upper():
                        txt_color = f"#{fc}"
                bold    = bool(cell.font and cell.font.bold)
                text    = format_cell_value_with_fmt(cell) if cell.value is not None else ""
                cell_w  = x2 - x1
                ch_w    = 8 if bold else 7
                max_chars = max(1, (cell_w - 8) // ch_w)
                # Detect header row
                if r == 1:
                # DON'T truncate header
                      display_text = text
                else:
                      display_text = text[:max_chars] if len(text) > max_chars else text

                      draw.text((x1 + 4, y1 + 4), display_text, fill=txt_color)
                draw.text((x1 + 4, y1 + 4), text, fill=txt_color)

    wb.close()
    return img, col_starts, row_starts, merged_master


def get_cell_pixel_bbox(
    col_starts: list, row_starts: list,
    target_row: int, target_col: int,
    merged_master: dict | None = None,
) -> tuple:
    c = max(1, min(target_col, len(col_starts) - 1))
    r = max(1, min(target_row, len(row_starts) - 1))
    if merged_master:
        info = merged_master.get((r, c))
        if info:
            mn_r, mn_c, mx_r, mx_c = info
            return (
                col_starts[mn_c - 1], row_starts[mn_r - 1],
                col_starts[min(mx_c, len(col_starts) - 1)],
                row_starts[min(mx_r, len(row_starts) - 1)],
            )
    return (
        col_starts[c - 1], row_starts[r - 1],
        col_starts[min(c, len(col_starts) - 1)],
        row_starts[min(r, len(row_starts) - 1)],
    )


def crop_context(img, x1, y1, x2, y2, pad_x: int = 220, pad_y: int = 160):
    iw, ih  = img.size
    cx1, cy1 = max(0, x1 - pad_x), max(0, y1 - pad_y)
    cx2, cy2 = min(iw, x2 + pad_x), min(ih, y2 + pad_y)
    return img.crop((cx1, cy1, cx2, cy2)), x1 - cx1, y1 - cy1, x2 - cx1, y2 - cy1

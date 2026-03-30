"""
modules/file_utils.py
Excel / CSV helpers: sheet names, dimensions, merged-cell metadata, totals rows.
"""

import csv
import os

import openpyxl
from openpyxl.utils import get_column_letter

from modules.cell_format import format_cell_value_with_fmt


# ── Sheet enumeration ─────────────────────────────────────────────────────────

def get_sheet_names(file_path: str) -> list[str]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return ["Sheet1"]
    wb      = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    names   = list(wb.sheetnames)
    wb.close()
    summary = [n for n in names if n.strip().lower() == "summary"]
    others  = [n for n in names if n.strip().lower() != "summary"]
    return summary + others


# ── Dimensions ────────────────────────────────────────────────────────────────

def get_sheet_dimensions(file_path: str, sheet_name: str) -> tuple[int, int]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        return len(rows), max((len(r) for r in rows), default=0)
    wb    = openpyxl.load_workbook(file_path, data_only=True)
    ws    = wb[sheet_name]
    max_r = ws.max_row or 0
    max_c = ws.max_column or 0
    if max_r == 0 or max_c == 0:
        actual_rows = actual_cols = 0
        for row in ws.iter_rows():
            if any(cell.value is not None for cell in row):
                actual_rows += 1
                row_col = max(
                    (cell.column for cell in row if cell.value is not None), default=0
                )
                actual_cols = max(actual_cols, row_col)
        max_r, max_c = actual_rows, actual_cols
    wb.close()
    return max_r, max_c


# ── Merged-cell metadata ──────────────────────────────────────────────────────

def extract_merged_cell_metadata(file_path: str, sheet_name: str) -> dict:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return {}
    wb          = openpyxl.load_workbook(file_path, data_only=True)
    ws          = wb[sheet_name]
    merged_info = {}
    for mr in ws.merged_cells.ranges:
        mn_r, mn_c, mx_r, mx_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        cell       = ws.cell(mn_r, mn_c)
        val        = str(cell.value).strip() if cell.value else ""
        span_cols  = mx_c - mn_c + 1
        span_rows  = mx_r - mn_r + 1
        region_type = (
            "TITLE"  if mn_r <= 3 and span_cols >= 3 else
            "HEADER" if span_cols >= 2 and span_rows == 1 else
            "DATA"
        )
        merged_info[f"R{mn_r}C{mn_c}"] = {
            "value": val, "type": region_type,
            "row_start": mn_r, "col_start": mn_c,
            "row_end": mx_r,   "col_end":   mx_c,
            "span_cols": span_cols, "span_rows": span_rows,
            "excel_row": mn_r,      "excel_col": mn_c,
        }
    wb.close()
    return merged_info


# ── Totals-row extraction ─────────────────────────────────────────────────────

def extract_totals_row(file_path: str, sheet_name: str) -> dict:
    ext    = os.path.splitext(file_path)[1].lower()
    totals = {}
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        cell_rows = None
    else:
        wb        = openpyxl.load_workbook(file_path, data_only=True)
        ws        = wb[sheet_name]
        raw_rows  = [[cell.value for cell in row] for row in ws.iter_rows()]
        cell_rows = [list(row) for row in ws.iter_rows()]
        rows      = raw_rows
        wb.close()
    if not rows:
        return totals

    header_row_index, headers = None, []
    for i, row in enumerate(rows[:20]):
        row_text = " ".join([str(c).lower() for c in row if c])
        if "claim" in row_text and ("date" in row_text or "incurred" in row_text or "paid" in row_text):
            header_row_index = i
            headers = [str(h).strip() if h is not None else f"Column_{j}" for j, h in enumerate(row)]
            break
    if header_row_index is None or not headers:
        return totals

    totals_rows = []
    for r_idx_rel, raw_row in enumerate(rows[header_row_index + 1:]):
        r_idx    = header_row_index + 2 + r_idx_rel
        if not any(raw_row):
            continue
        row_text = " ".join([str(c).lower() for c in raw_row if c])
        if any(kw in row_text for kw in ["total", "subtotal", "grand total", "sum", "totals"]):
            row_data: dict = {}
            cell_row = cell_rows[header_row_index + 1 + r_idx_rel] if cell_rows else None
            for c_idx_0, raw_val in enumerate(raw_row):
                if c_idx_0 >= len(headers):
                    continue
                if cell_row and c_idx_0 < len(cell_row):
                    clean_val = format_cell_value_with_fmt(cell_row[c_idx_0])
                    real_col  = cell_row[c_idx_0].column if hasattr(cell_row[c_idx_0], "column") else c_idx_0 + 1
                else:
                    clean_val = str(raw_val).strip() if raw_val is not None else ""
                    real_col  = c_idx_0 + 1
                if clean_val:
                    row_data[headers[c_idx_0]] = {
                        "value": clean_val, "excel_row": r_idx, "excel_col": real_col,
                    }
            if row_data:
                totals_rows.append(row_data)

    if totals_rows:
        totals["rows"]      = totals_rows
        totals["excel_row"] = totals_rows[0].get(list(totals_rows[0].keys())[0], {}).get("excel_row", 9999)
        agg: dict = {}
        for row_data in totals_rows:
            for field, info in row_data.items():
                try:
                    num = float(str(info["value"]).replace(",", "").replace("$", ""))
                    agg[field] = agg.get(field, 0.0) + num
                except Exception:
                    pass
        totals["aggregated"] = {k: round(v, 2) for k, v in agg.items()}
    return totals


# ── Compute totals from parsed claim data (always available) ──────────────────

def compute_totals_from_claims(claims_data: list[dict]) -> dict:
    """
    Calculate column totals directly from the parsed claim rows.

    Each item in `claims_data` is a dict of  { field_name: {"value": ..., ...} }.
    Only numeric-looking values are summed.  Non-numeric columns are skipped.

    Returns a dict shaped like extract_totals_row():
        {
            "aggregated": { field: rounded_float, ... },
            "source":     "computed",          # marks that this was not from an Excel totals row
        }
    Always returns at least {"aggregated": {}, "source": "computed"} — never an
    empty dict — so callers can rely on the "aggregated" key existing.
    """
    if not claims_data:
        return {"aggregated": {}, "source": "computed"}

    agg: dict[str, float] = {}

    for claim in claims_data:
        for field, info in claim.items():
            raw = info.get("modified") or info.get("value", "")
            if raw is None:
                continue
            cleaned = str(raw).strip().replace(",", "").replace("$", "").replace("%", "")
            try:
                num = float(cleaned)
            except (ValueError, TypeError):
                continue
            agg[field] = round(agg.get(field, 0.0) + num, 2)

    return {"aggregated": agg, "source": "computed"}


def get_totals_for_sheet(
    file_path: str,
    sheet_name: str,
    claims_data: list[dict],
) -> dict:
    """
    Always return a populated totals dict for the sheet.

    Strategy:
      1. Try to find an explicit totals row in the Excel file (existing behaviour).
      2. If none found (or aggregated is empty), fall back to computing totals
         directly from `claims_data`.

    The returned dict always contains an "aggregated" key.
    """
    totals = extract_totals_row(file_path, sheet_name)

    # Use the Excel totals row only when it actually produced numbers
    if totals.get("aggregated"):
        totals.setdefault("source", "excel_row")
        return totals

    # Fall back: compute from the parsed claim rows
    return compute_totals_from_claims(claims_data)
